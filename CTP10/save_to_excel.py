import os  # 파일 경로 처리, 존재 여부 확인 등에 사용하는 표준 라이브러리
import pandas as pd  # 엑셀 쓰기/읽기 및 DataFrame 조작을 위한 pandas
from openpyxl import Workbook  # openpyxl의 워크북 객체 (직접 생성 시 사용)
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment  # 셀 스타일(채우기, 폰트, 테두리, 정렬) 지정

def save_to_excel(results_df, output_excel_file_path, rois):  # 비교 결과 DataFrame을 엑셀로 저장하고 서식을 입히는 함수
    with pd.ExcelWriter(output_excel_file_path, engine='openpyxl') as writer:  # openpyxl 엔진으로 엑셀 작성 컨텍스트 열기
        results_df.to_excel(writer, index=False, sheet_name='Results')  # DataFrame을 'Results' 시트로 저장(인덱스 제외)
        workbook = writer.book  # 작성 중인 openpyxl Workbook 객체 참조
        worksheet = writer.sheets['Results']  # 방금 쓴 'Results' 워크시트 객체 가져오기

        # 첫 번째 열을 고정 (B2 기준으로 고정: 1행과 A열이 고정됨)
        worksheet.freeze_panes = worksheet['B2']  # 'B2'를 기준으로 위/왼쪽을 고정하여 스크롤 시 헤더/첫 열 고정

        # Patient ID 열의 인덱스 찾기 (DataFrame은 0-base이므로 openpyxl의 1-base 컬럼 인덱스 위해 +1)
        patient_id_col_idx = results_df.columns.get_loc('Patient ID') + 1  # 'Patient ID'가 몇 번째 컬럼인지 계산

        # 초록색으로 컬러 설정 (주의: openpyxl의 색상 코드는 일반적으로 RRGGBB 또는 ARGB 사용)
        green_fill = PatternFill(start_color='D0EA00', end_color='00FF00', fill_type='solid')  # Patient ID 셀 배경 채우기 색
        pass_font = Font(color='00C300', bold=True)  # 'Pass' 텍스트에 적용할 녹색 굵은 폰트
        fail_font = Font(color='FF0000', bold=True)  # 'Fail' 텍스트에 적용할 빨간 굵은 폰트
        bold_side = Side(border_style='medium', color='000000')  # 테두리용 선 정의(중간 두께, 검정)

        for row in range(2, len(results_df) + 2):  # 헤더는 1행이므로 실제 데이터는 2행부터 시작, 마지막 행까지 순회
            cell = worksheet.cell(row=row, column=patient_id_col_idx)  # 현재 행의 'Patient ID' 셀 객체
            cell.fill = green_fill  # Patient ID 셀에 초록 배경색 채우기 적용

            # 전체 결과(마지막 컬럼이 Overall/Result라고 가정) 컬럼에 대한 스타일 적용
            overall_result_cell = worksheet.cell(row=row, column=len(results_df.columns))  # 마지막 컬럼 셀 참조
            if overall_result_cell.value == 'Pass':  # 셀 값이 'Pass'라면
                overall_result_cell.font = pass_font  # 녹색 굵은 폰트 적용
            elif overall_result_cell.value == 'Fail':  # 셀 값이 'Fail'라면
                overall_result_cell.font = fail_font  # 빨간 굵은 폰트 적용

            # 개별 ROI 결과 및 수치 컬럼들에 대한 스타일 적용 (Patient ID 다음 컬럼부터 끝까지)
            for col in range(patient_id_col_idx + 1, len(results_df.columns) + 1):  # openpyxl은 1-base 인덱스
                cell = worksheet.cell(row=row, column=col)  # 현재 셀 참조
                cell.alignment = Alignment(horizontal='right')  # 숫자/결과의 기본 정렬을 우측 정렬로 설정
                if cell.value == 'Pass':  # 셀 값이 'Pass'면
                    cell.font = pass_font  # 녹색 굵은 폰트
                elif cell.value == 'Fail':  # 셀 값이 'Fail'면
                    cell.font = fail_font  # 빨간 굵은 폰트
                elif results_df.columns[col-1].endswith('system'):  # 현재 컬럼명이 '... system'으로 끝나면 (메타/시스템 값)
                    cell.font = Font(bold=True)  # 해당 시스템 컬럼은 굵게 처리 (색상은 기본값 유지)

        # ROI 그룹별로 테두리 적용 (각 ROI에 대해 Result/min/system/max/Differ 5개 컬럼 묶음에 외곽선)
        for roi in rois:  # 전달받은 ROI 이름 리스트 순회
            result_col = results_df.columns.get_loc(f'{roi} Result') + 1  # '{ROI} Result' 컬럼의 1-base 위치
            min_col = results_df.columns.get_loc(f'{roi} min') + 1  # '{ROI} min' 컬럼 위치
            system_col = results_df.columns.get_loc(f'{roi} system') + 1  # '{ROI} system' 컬럼 위치
            max_col = results_df.columns.get_loc(f'{roi} max') + 1  # '{ROI} max' 컬럼 위치
            differ_col = results_df.columns.get_loc(f'{roi} Differ') + 1  # '{ROI} Differ' 컬럼 위치(신규 추가 컬럼)

            for row in range(2, len(results_df) + 2):  # 데이터 영역 모든 행에 대해
                worksheet.cell(row=row, column=result_col).border = Border(left=bold_side, top=bold_side, bottom=bold_side)  # 묶음의 왼쪽 경계(RESULT)에 좌/상/하 테두리
                worksheet.cell(row=row, column=min_col).border = Border(top=bold_side, bottom=bold_side)  # MIN은 상/하 테두리
                worksheet.cell(row=row, column=system_col).border = Border(top=bold_side, bottom=bold_side)  # SYSTEM도 상/하 테두리
                worksheet.cell(row=row, column=max_col).border = Border(top=bold_side, bottom=bold_side)  # MAX도 상/하 테두리
                worksheet.cell(row=row, column=differ_col).border = Border(right=bold_side, top=bold_side, bottom=bold_side)  # 묶음의 오른쪽 경계(DIFFER)에 우/상/하 테두리
