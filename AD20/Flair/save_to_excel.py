import pandas as pd
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

def save_to_excel(results_df, output_excel_file_path, rois):
    # Overall Result 열 이동
    if 'Overall Result' in results_df.columns:
        overall_col = results_df.pop('Overall Result')
        insert_position = results_df.columns.get_loc('Patient ID') + 1
        results_df.insert(insert_position, 'Overall Result', overall_col)
    else:
        insert_position = results_df.columns.get_loc('Patient ID') + 1  # 안전용

    with pd.ExcelWriter(output_excel_file_path, engine='openpyxl') as writer:
        # 1) 먼저 쓰기
        results_df.to_excel(writer, index=False, sheet_name='Results')
        workbook  = writer.book
        worksheet = writer.sheets['Results']

        # 2) 서식/고정 모두 with 안에서 처리 (→ 실제 파일에 저장됨)
        # 첫 번째 열 고정
        worksheet.freeze_panes = worksheet['B2']

        # 인덱스
        patient_id_col_idx = results_df.columns.get_loc('Patient ID') + 1

        # 색/폰트/테두리 (ARGB 8자리 사용)
        green_fill = PatternFill(fill_type='solid', fgColor='FFD0EA00')  # 밝은 연두색
        pass_font  = Font(color='FF00C300', bold=True)  # 초록
        fail_font  = Font(color='FFFF0000', bold=True)  # 빨강
        bold_side  = Side(border_style='medium', color='FF000000')

        # 행 서식
        for row in range(2, len(results_df) + 2):  # 헤더 제외
            # Patient ID 배경
            worksheet.cell(row=row, column=patient_id_col_idx).fill = green_fill

            # Overall Result 서식 (있을 때만)
            if 'Overall Result' in results_df.columns:
                overall_col_idx = results_df.columns.get_loc('Overall Result') + 1
                overall_cell = worksheet.cell(row=row, column=overall_col_idx)
                if overall_cell.value == 'Pass':
                    overall_cell.font = pass_font
                elif overall_cell.value == 'Fail':
                    overall_cell.font = fail_font

            # 나머지 셀 정렬/Pass/Fail/볼드
            for col in range(patient_id_col_idx + 1, len(results_df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='right')
                if cell.value == 'Pass':
                    cell.font = pass_font
                elif cell.value == 'Fail':
                    cell.font = fail_font
                elif str(results_df.columns[col-1]).endswith('system'):
                    cell.font = Font(bold=True)

        # ROI 블록 테두리 (존재하는 컬럼에만)
        for roi in rois:
            try:
                result_col = results_df.columns.get_loc(f'{roi} Result') + 1
                min_col    = results_df.columns.get_loc(f'{roi} min') + 1
                system_col = results_df.columns.get_loc(f'{roi} system') + 1
                max_col    = results_df.columns.get_loc(f'{roi} max') + 1
                differ_col = results_df.columns.get_loc(f'{roi} Differ') + 1

                for row in range(2, len(results_df) + 2):
                    worksheet.cell(row=row, column=result_col).border = Border(left=bold_side, top=bold_side, bottom=bold_side)
                    worksheet.cell(row=row, column=min_col).border    = Border(top=bold_side, bottom=bold_side)
                    worksheet.cell(row=row, column=system_col).border = Border(top=bold_side, bottom=bold_side)
                    worksheet.cell(row=row, column=max_col).border    = Border(top=bold_side, bottom=bold_side)
                    worksheet.cell(row=row, column=differ_col).border = Border(right=bold_side, top=bold_side, bottom=bold_side)
            except KeyError:
                # 특정 ROI 블록 컬럼이 없으면 스킵 (필요시 로그)
                continue

        # with 종료 시점에 writer가 실제 파일 저장
