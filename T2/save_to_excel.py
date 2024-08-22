import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment

def save_to_excel(results_df, output_excel_file_path, rois):
    # 'Overall Result' 컬럼을 원하는 위치로 이동 (예: Patient ID 뒤에 위치시키기)
    if 'Overall Result' in results_df.columns:
        overall_col = results_df.pop('Overall Result')
        insert_position = results_df.columns.get_loc('Patient ID') + 1
        results_df.insert(insert_position, 'Overall Result', overall_col)

    with pd.ExcelWriter(output_excel_file_path, engine='openpyxl') as writer:
        results_df.to_excel(writer, index=False, sheet_name='Results')
        workbook = writer.book
        worksheet = writer.sheets['Results']

        # 첫 번째 열을 고정
        worksheet.freeze_panes = worksheet['B2']

        # Patient ID 열의 인덱스 찾기
        patient_id_col_idx = results_df.columns.get_loc('Patient ID') + 1

        # 초록색으로 컬러 설정
        green_fill = PatternFill(start_color='D0EA00', end_color='00FF00', fill_type='solid')
        pass_font = Font(color='00C300', bold=True)
        fail_font = Font(color='FF0000', bold=True)
        bold_side = Side(border_style='medium', color='000000')

        for row in range(2, len(results_df) + 2):  # 헤더 이후부터 시작
            cell = worksheet.cell(row=row, column=patient_id_col_idx)
            cell.fill = green_fill

            # 전체 결과 컬럼에 대한 스타일 적용
            overall_result_cell = worksheet.cell(row=row, column=insert_position + 1)
            if overall_result_cell.value == 'Pass':
                overall_result_cell.font = pass_font
            elif overall_result_cell.value == 'Fail':
                overall_result_cell.font = fail_font

            # 개별 ROI 결과 컬럼에 대한 스타일 적용
            for col in range(patient_id_col_idx + 1, len(results_df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal='right')
                if cell.value == 'Pass':
                    cell.font = pass_font
                elif cell.value == 'Fail':
                    cell.font = fail_font
                elif results_df.columns[col-1].endswith('system'):  # 'system' 열인 경우
                    cell.font = Font(bold=True)

        # ROI 그룹별로 테두리 적용
        for roi in rois:
            result_col = results_df.columns.get_loc(f'{roi} Result') + 1
            min_col = results_df.columns.get_loc(f'{roi} min') + 1
            system_col = results_df.columns.get_loc(f'{roi} system') + 1
            max_col = results_df.columns.get_loc(f'{roi} max') + 1
            differ_col = results_df.columns.get_loc(f'{roi} Differ') + 1  # 새로 추가된 Differ 열 위치

            for row in range(2, len(results_df) + 2):
                worksheet.cell(row=row, column=result_col).border = Border(left=bold_side, top=bold_side, bottom=bold_side)
                worksheet.cell(row=row, column=min_col).border = Border(top=bold_side, bottom=bold_side)
                worksheet.cell(row=row, column=system_col).border = Border(top=bold_side, bottom=bold_side)
                worksheet.cell(row=row, column=max_col).border = Border(top=bold_side, bottom=bold_side)
                worksheet.cell(row=row, column=differ_col).border = Border(right=bold_side, top=bold_side, bottom=bold_side)
