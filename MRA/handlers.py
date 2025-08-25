# SCALE/MRA/handlers.py
import io
import pandas as pd
from typing import Iterable, Dict, Tuple, List, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# 비교 대상 컬럼 (정답지의 roi_product_name 값과 매칭)
COMPARE_COLS: Iterable[str] = (
    "CoordinateX",
    "CoordinateY",
    "CoordinateZ",
    "Maximum Diameter",
    "Probability Score (RUO)",
    "Location (RUO)",
)

# Location 코드 → 라벨 매핑
LOCATION_MAP: Dict[int, str] = {
    1: "PICA, left", 2: "PICA, right", 3: "AICA, left", 4: "AICA, right",
    5: "SCA, left", 6: "SCA, right", 7: "BA",
    8: "communicating ICA, left", 9: "communicating ICA, right",
    10: "communicating ICA, left", 11: "communicating ICA, right",
    12: "cavernous ICA, left", 13: "cavernous ICA, right",
    14: "clinoid-ophthalmic ICA, left", 15: "clinoid-ophthalmic ICA, right",
    16: "clinoid-ophthalmic ICA, left", 17: "clinoid-ophthalmic ICA, right",
    18: "clinoid-ophthalmic ICA, left", 19: "clinoid-ophthalmic ICA, right",
    20: "clinoid-ophthalmic ICA, left", 21: "clinoid-ophthalmic ICA, right",
    22: "communicating ICA, left", 23: "communicating ICA, right",
    24: "MCA, left", 25: "MCA, right", 26: "MCA, left", 27: "MCA, right",
    28: "ACOM", 29: "distal ACA, left", 30: "distal ACA, right",
}

def _fx(val):
    try:
        if pd.isna(val):
            return None
        return float(val)
    except Exception:
        return None

# ---------------- 1) read_data (파일 경로 기반) ----------------
def read_data(csv_path: str, excel_path: str) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    process_comparison 1단계: 파일 경로에서 DataFrame 로드
    반환: (csv_df, excel_df_filtered)
    """
    csv_df = pd.read_csv(csv_path)
    excel_df = pd.read_excel(excel_path)
    excel_df.columns = excel_df.columns.str.strip()

    # 비교대상만 필터링 (불필요한 정답지 행 제거)
    excel_df_filtered = excel_df[excel_df.get("roi_product_name", "").isin(COMPARE_COLS)].copy()

    # 정답지 필수 컬럼 검증(초기 검증은 여기서 한번)
    required_xlsx = {
        "session_id","Aneurysm Index","roi_product_name",
        "Engine_raw_vol_min","Engine_raw_vol_max","Engine_raw_vol_mean","Diameter",
    }
    missing = required_xlsx - set(excel_df.columns)
    if missing:
        raise ValueError(f"XLSX missing columns: {sorted(missing)}")

    # CSV 필수 컬럼 검증
    required_csv = {"Patient ID","Aneurysm Index"}
    missing_csv = required_csv - set(csv_df.columns)
    if missing_csv:
        raise ValueError(f"CSV missing columns: {sorted(missing_csv)}")

    return csv_df, excel_df_filtered

# ---------------- 2) compare_data ----------------
def compare_data(csv_df: pd.DataFrame, excel_df_filtered: pd.DataFrame) -> Tuple[pd.DataFrame, List[str]]:
    """
    process_comparison 2단계: 두 DataFrame 비교
    반환: (results_df, rois)
      - rois: 결과 엑셀 스타일링 등에 쓰일 ROI 이름 리스트(옵션)
    """
    results: List[Dict[str, Any]] = []
    rois: List[str] = list(sorted(set(excel_df_filtered["roi_product_name"].tolist())))

    keys = excel_df_filtered[["session_id", "Aneurysm Index"]].drop_duplicates()

    for _, k in keys.iterrows():
        sid = k["session_id"]
        idx = k["Aneurysm Index"]

        ref_group = excel_df_filtered[
            (excel_df_filtered["session_id"] == sid) &
            (excel_df_filtered["Aneurysm Index"] == idx)
        ]

        match_row = csv_df[
            (csv_df["Patient ID"] == sid) &
            (csv_df["Aneurysm Index"] == idx)
        ]

        if match_row.empty:
            for _, r in ref_group.iterrows():
                results.append({
                    "Patient ID": sid,
                    "Aneurysm Index": idx,
                    "ROI Name": r["roi_product_name"],
                    "Value (CSV)": None,
                    "Min (XLSX)": r["Engine_raw_vol_min"],
                    "Max (XLSX)": r["Engine_raw_vol_max"],
                    "Result": "NoMatch",
                })
            continue

        row = match_row.iloc[0]  # 동일 키 다건이면 첫 행 사용

        for _, r in ref_group.iterrows():
            feat = r["roi_product_name"]

            if feat == "Location (RUO)":
                raw = row.get(feat, None)
                if isinstance(raw, str):
                    val_csv = raw
                else:
                    val_csv = LOCATION_MAP.get(raw, None)
                val_ref = r["Engine_raw_vol_mean"]

                if val_csv is None or pd.isna(val_ref):
                    result = "NoMatch"
                elif str(val_csv).strip().lower() == str(val_ref).strip().lower():
                    result = "Pass"
                else:
                    result = "Fail"

                results.append({
                    "Patient ID": sid, "Aneurysm Index": idx,
                    "ROI Name": feat,
                    "Value (CSV)": val_csv,
                    "Min (XLSX)": val_ref,
                    "Max (XLSX)": None,
                    "Result": result,
                })
                continue

            if feat == "Maximum Diameter":
                f_csv = _fx(row.get(feat, None))
                f_ref = _fx(r.get("Diameter", None))
                if f_csv is None or f_ref is None:
                    result = "NoMatch"
                elif f_csv == f_ref:
                    result = "Pass"
                else:
                    result = "Fail"

                results.append({
                    "Patient ID": sid, "Aneurysm Index": idx,
                    "ROI Name": feat,
                    "Value (CSV)": row.get(feat, None),
                    "Min (XLSX)": r.get("Diameter", None),
                    "Max (XLSX)": None,
                    "Result": result,
                })
                continue

            # 기타 수치형: min~max 범위 비교
            f_csv = _fx(row.get(feat, None))
            f_min = _fx(r["Engine_raw_vol_min"])
            f_max = _fx(r["Engine_raw_vol_max"])

            if f_csv is None or f_min is None or f_max is None:
                result = "NoMatch"
            elif f_min <= f_csv <= f_max:
                result = "Pass"
            else:
                result = "Fail"

            results.append({
                "Patient ID": sid, "Aneurysm Index": idx,
                "ROI Name": feat,
                "Value (CSV)": row.get(feat, None),
                "Min (XLSX)": r["Engine_raw_vol_min"],
                "Max (XLSX)": r["Engine_raw_vol_max"],
                "Result": result,
            })

    results_df = pd.DataFrame(results)
    return results_df, rois

# ---------------- 3) save_to_excel (파일 경로 기반) ----------------
def save_to_excel(results_df: pd.DataFrame, output_excel_file_path: str, rois: Iterable[str]) -> None:
    """
    process_comparison 3단계: 결과 엑셀 파일로 저장(스타일 포함)
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "비교 결과"

    fill_pass = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")
    fill_fail = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    bold = Font(bold=True)

    for r_idx, row in enumerate(dataframe_to_rows(results_df, index=False, header=True), start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if r_idx == 1:
                cell.font = bold
            else:
                if c_idx == len(row):  # 마지막 컬럼(Result)
                    cell.font = bold
                    if val == "Pass":
                        cell.fill = fill_pass
                    elif val == "Fail":
                        cell.fill = fill_fail

    wb.save(output_excel_file_path)
