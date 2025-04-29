import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows

# 1. 파일 불러오기
df_csv = pd.read_csv('SCALE MRA_Results.csv')
df_xlsx = pd.read_excel('MRA_Answer.xlsx')
df_xlsx.columns = df_xlsx.columns.str.strip()  # 공백 제거

# 2. 비교 대상 컬럼 정의
compare_cols = [
    'CoordinateX', 'CoordinateY', 'CoordinateZ',
    'Maximum Diameter', 'Probability Score (RUO)', 'Location (RUO)'
]

# 3. Location 매핑
location_map = {
    1: 'PICA, left', 2: 'PICA, right', 3: 'AICA, left', 4: 'AICA, right',
    5: 'SCA, left', 6: 'SCA, right', 7: 'BA',
    8: 'communicating ICA, left', 9: 'communicating ICA, right',
    10: 'communicating ICA, left', 11: 'communicating ICA, right',
    12: 'cavernous ICA, left', 13: 'cavernous ICA, right',
    14: 'clinoid-ophthalmic ICA, left', 15: 'clinoid-ophthalmic ICA, right',
    16: 'clinoid-ophthalmic ICA, left', 17: 'clinoid-ophthalmic ICA, right',
    18: 'clinoid-ophthalmic ICA, left', 19: 'clinoid-ophthalmic ICA, right',
    20: 'clinoid-ophthalmic ICA, left', 21: 'clinoid-ophthalmic ICA, right',
    22: 'communicating ICA, left', 23: 'communicating ICA, right',
    24: 'MCA, left', 25: 'MCA, right', 26: 'MCA, left', 27: 'MCA, right',
    28: 'ACOM', 29: 'distal ACA, left', 30: 'distal ACA, right'
}

# 4. 기준 데이터 정리 (Diameter 컬럼 포함)
df_ref = df_xlsx[df_xlsx['roi_product_name'].isin(compare_cols)].copy()
df_ref = df_ref[['session_id', 'Aneurysm Index', 'roi_product_name',
                 'Engine_raw_vol_min', 'Engine_raw_vol_max', 'Engine_raw_vol_mean', 'Diameter']]

# 5. 결과 저장 리스트
results = []

# 6. 기준 데이터에 있는 모든 session_id + Aneurysm Index 쌍에 대해 비교 수행
unique_keys = df_ref[['session_id', 'Aneurysm Index']].drop_duplicates()

for _, key in unique_keys.iterrows():
    session_id = key['session_id']
    aneurysm_idx = key['Aneurysm Index']

    ref_group = df_ref[
        (df_ref['session_id'] == session_id) &
        (df_ref['Aneurysm Index'] == aneurysm_idx)
    ]

    match_row = df_csv[
        (df_csv['Patient ID'] == session_id) &
        (df_csv['Aneurysm Index'] == aneurysm_idx)
    ]

    if match_row.empty:
        for _, r in ref_group.iterrows():
            results.append({
                'Patient ID': session_id,
                'Aneurysm Index': aneurysm_idx,
                'ROI Name': r['roi_product_name'],
                'Value (CSV)': None,
                'Min (XLSX)': r['Engine_raw_vol_min'],
                'Max (XLSX)': r['Engine_raw_vol_max'],
                'Result': 'NoMatch'
            })
    else:
        row = match_row.iloc[0]
        for _, r in ref_group.iterrows():
            feature = r['roi_product_name']

            # --- Location 비교 (문자열 매칭) ---
            if feature == 'Location (RUO)':
                loc_code = row.get(feature, None)
                val_csv = location_map.get(loc_code, None)
                val_ref = r['Engine_raw_vol_mean']

                if val_csv is None or pd.isna(val_ref):
                    result = 'NoMatch'
                elif val_csv.strip().lower() == val_ref.strip().lower():
                    result = 'Pass'
                else:
                    result = 'Fail'

                results.append({
                    'Patient ID': session_id,
                    'Aneurysm Index': aneurysm_idx,
                    'ROI Name': feature,
                    'Value (CSV)': val_csv,
                    'Min (XLSX)': val_ref,
                    'Max (XLSX)': None,
                    'Result': result
                })

            # --- Maximum Diameter 비교 (정확히 일치) ---
            elif feature == 'Maximum Diameter':
                val_csv = row.get(feature, None)
                val_ref = r.get('Diameter', None)

                try:
                    if pd.isna(val_csv) or pd.isna(val_ref):
                        result = 'NoMatch'
                    elif float(val_csv) == float(val_ref):
                        result = 'Pass'
                    else:
                        result = 'Fail'
                except:
                    result = 'NoMatch'

                results.append({
                    'Patient ID': session_id,
                    'Aneurysm Index': aneurysm_idx,
                    'ROI Name': feature,
                    'Value (CSV)': val_csv,
                    'Min (XLSX)': val_ref,
                    'Max (XLSX)': None,
                    'Result': result
                })

            # --- 나머지 수치형 항목은 범위 비교 ---
            else:
                val_csv = row.get(feature, None)
                min_val = r['Engine_raw_vol_min']
                max_val = r['Engine_raw_vol_max']

                try:
                    if pd.isna(min_val) or pd.isna(max_val) or pd.isna(val_csv):
                        result = 'NoMatch'
                    elif float(min_val) <= float(val_csv) <= float(max_val):
                        result = 'Pass'
                    else:
                        result = 'Fail'
                except:
                    result = 'NoMatch'

                results.append({
                    'Patient ID': session_id,
                    'Aneurysm Index': aneurysm_idx,
                    'ROI Name': feature,
                    'Value (CSV)': val_csv,
                    'Min (XLSX)': min_val,
                    'Max (XLSX)': max_val,
                    'Result': result
                })

# 7. 결과 DataFrame 생성
df_result = pd.DataFrame(results)

# 8. 엑셀 스타일링
wb = Workbook()
ws = wb.active
ws.title = "비교 결과"

fill_pass = PatternFill(start_color='CCFFCC', end_color='CCFFCC', fill_type='solid')
fill_fail = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
bold_font = Font(bold=True)

for r_idx, row in enumerate(dataframe_to_rows(df_result, index=False, header=True), start=1):
    for c_idx, val in enumerate(row, start=1):
        cell = ws.cell(row=r_idx, column=c_idx, value=val)
        if r_idx == 1:
            cell.font = bold_font
        elif c_idx == len(row):
            cell.font = bold_font
            if val == 'Pass':
                cell.fill = fill_pass
            elif val == 'Fail':
                cell.fill = fill_fail

# 9. 엑셀 저장
wb.save('MRA_Validation.xlsx')
print("✅ 스타일이 적용된 최종 결과가 MRA_Validation.xlsx 로 저장되었습니다.")
